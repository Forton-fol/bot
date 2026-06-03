from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from repositories.game import GameRepository
from repositories.points import PointsRepository
from repositories.user import UserRepository


class ExcelExportService:
    def __init__(self, session: AsyncSession) -> None:
        self.session = session
        self.users = UserRepository(session)
        self.games = GameRepository(session)
        self.points = PointsRepository(session)

    async def export_statistics(self, output_dir: str = "/tmp") -> Path:
        stats = await self.games.get_statistics()
        users = await self.users.list_all()
        games = await self.games.list_all_for_admin(500)
        history = await self.points.get_all_for_export()

        wb = Workbook()

        ws_stats = wb.active
        ws_stats.title = "Статистика"
        ws_stats.append(["Показатель", "Значение"])
        for key, value in stats.items():
            ws_stats.append([key, value])

        ws_users = wb.create_sheet("Игроки")
        ws_users.append([
            "ID", "Telegram ID", "Имя", "Username", "Роль", "Баллы", "Посещений"
        ])
        for u in users:
            ws_users.append([
                u.id,
                u.telegram_id,
                u.full_name,
                u.username or "",
                u.role_name,
                u.points,
                u.games_attended,
            ])

        ws_games = wb.create_sheet("Игры")
        ws_games.append([
            "ID", "Название", "Дата", "Время", "Статус", "Мастер", "Мест", "Занято"
        ])
        for g in games:
            ws_games.append([
                g.id,
                g.title,
                str(g.game_date),
                str(g.game_time),
                g.status,
                g.master.full_name if g.master else "",
                g.max_slots,
                g.occupied_slots,
            ])

        ws_points = wb.create_sheet("Баллы")
        ws_points.append(["User ID", "Сумма", "Причина", "Игра ID", "Дата"])
        for h in history:
            ws_points.append([
                h.user_id,
                h.amount,
                h.reason,
                h.game_id or "",
                h.created_at.isoformat() if h.created_at else "",
            ])

        Path(output_dir).mkdir(parents=True, exist_ok=True)
        filename = f"nri_stats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path = Path(output_dir) / filename
        wb.save(path)
        return path
